"""
RPL Sports Prediction Challenge — Final Pipeline
==================================================
Ensemble of 5 rating systems to predict derby match win margins
and rank all 165 teams in the Rocketball Premier League.

Author: Maanas
Model: 5-system weighted ensemble with linear calibration + shrinkage
"""

import pandas as pd
import numpy as np
from copy import deepcopy
from numpy.polynomial.polynomial import polyfit
import openpyxl

# ════════════════════════════════════════════════════════════════════
# LOAD DATA
# ════════════════════════════════════════════════════════════════════

train = pd.read_csv('Train.csv')       # 940 regular season games
pred  = pd.read_csv('Predictions.csv')  # 75 derby matches (fill in margins)

# Build team index (165 teams)
all_teams = sorted(set(train['HomeTeam'].tolist() + train['AwayTeam'].tolist()))
team_to_idx = {t: i for i, t in enumerate(all_teams)}
n_teams = len(all_teams)

# Map team names → IDs (for rankings output)
team_to_id = {}
for _, row in train.iterrows():
    team_to_id[row['HomeTeam']] = row['HomeID']
    team_to_id[row['AwayTeam']] = row['AwayID']


# ════════════════════════════════════════════════════════════════════
# STEP 1: ESTIMATE HOME FIELD ADVANTAGE VIA REGRESSION
# ════════════════════════════════════════════════════════════════════
# Instead of using the naive mean margin (5.9 pts), we estimate HFA
# jointly with team ratings using ridge regression. This controls for
# the fact that strong teams may have played more home games, which
# biases the raw mean upward.
#
# Model: HomeWinMargin = rating[home] - rating[away] + HFA + noise
# Solve via ridge: [M | 1] @ [ratings; HFA] ≈ margins

M = np.zeros((len(train), n_teams))
y_raw = train['HomeWinMargin'].values.astype(float)

for i, row in train.iterrows():
    M[i, team_to_idx[row['HomeTeam']]]  =  1
    M[i, team_to_idx[row['AwayTeam']]]  = -1

M_hfa = np.column_stack([M, np.ones(len(train))])  # append intercept = HFA
lam = 0.5  # ridge penalty

sol = np.linalg.solve(
    M_hfa.T @ M_hfa + lam * np.eye(M_hfa.shape[1]),
    M_hfa.T @ y_raw
)
HFA = sol[n_teams]  # learned home field advantage = 3.31 pts
print(f"Regression-estimated Home Field Advantage: {HFA:.2f} pts")


# ════════════════════════════════════════════════════════════════════
# STEP 2: BUILD 5 INDEPENDENT RATING SYSTEMS
# ════════════════════════════════════════════════════════════════════
# Using diverse methods ensures the ensemble captures different
# aspects of team strength and is robust to any single method's flaws.


# ─── System 1: Massey Ratings (Ridge Regression) ───────────────────
# The Massey method frames rating estimation as a least-squares problem:
#   rating[home] - rating[away] ≈ margin   for every game
# We solve this system via ridge regression (λ=0.5) which also
# provides regularization to handle the underdetermined system.
# The HFA column lets us learn home advantage jointly with ratings.

massey_ratings = sol[:n_teams]
massey_ratings -= massey_ratings.mean()  # center at 0
massey_dict = {all_teams[i]: massey_ratings[i] for i in range(n_teams)}


# ─── System 2: Elo Ratings with Margin of Victory ─────────────────
# Elo is a sequential rating system: after each game, ratings update
# based on the surprise of the result. We extend standard Elo with:
#   - A margin-of-victory multiplier (log-scaled to dampen blowouts)
#   - An autocorrelation correction (prevents double-counting big upsets)
#   - A home advantage of 50 Elo points
# We run 5 full passes through the season to stabilize ratings.

K = 32              # base update magnitude
HOME_ADV_ELO = 50   # Elo points for home advantage

train_sorted = train.sort_values('Date').reset_index(drop=True)

for _ in range(5):  # multiple passes for convergence
    elo = {t: 1500.0 for t in all_teams}
    for _, row in train_sorted.iterrows():
        home, away = row['HomeTeam'], row['AwayTeam']
        margin = row['HomeWinMargin']

        # Expected win probability for home team
        elo_diff = elo[home] + HOME_ADV_ELO - elo[away]
        expected = 1 / (1 + 10 ** (-elo_diff / 400))

        # MOV multiplier: log-scaled margin, dampened by rating gap
        # This prevents strong favorites from getting huge boosts for expected blowouts
        mov_mult = np.log(abs(margin) + 1) * (2.2 / (abs(elo_diff) * 0.001 + 2.2))

        actual = 1.0 if margin > 0 else (0.5 if margin == 0 else 0.0)
        update = K * mov_mult * (actual - expected)
        elo[home] += update
        elo[away] -= update

elo_ratings = {t: r - 1500 for t, r in elo.items()}


# ─── System 3: Adjusted Point Differential with SOS ───────────────
# Raw point differential is the simplest team strength measure, but
# it doesn't account for schedule difficulty. We iteratively adjust:
#   adjusted_rating[t] = raw_margin[t] + 0.3 × mean(adjusted_rating[opponents])
# After 3 iterations, each team's rating reflects both their own
# performance AND the quality of teams they played against.

team_margins = {t: [] for t in all_teams}
team_opps    = {t: [] for t in all_teams}

for _, row in train.iterrows():
    h, a = row['HomeTeam'], row['AwayTeam']
    neutral_margin = row['HomeWinMargin'] - HFA  # remove home advantage
    team_margins[h].append(neutral_margin)
    team_margins[a].append(-neutral_margin)
    team_opps[h].append(a)
    team_opps[a].append(h)

raw_avg = {t: np.mean(ms) for t, ms in team_margins.items()}
adj_rating = deepcopy(raw_avg)

for _ in range(3):  # 3 SOS adjustment iterations
    adj_rating = {
        t: raw_avg[t] + 0.3 * np.mean([adj_rating[o] for o in team_opps[t]])
        for t in all_teams
    }


# ─── System 4: Colley Matrix with Fractional Wins ─────────────────
# The Colley method starts from a Bayesian prior (each team is 0.500)
# and updates based on win/loss records. We modify it to use
# "fractional wins" via a sigmoid function on the margin:
#   frac_win = sigmoid(margin / 20)
# This means a 40-point blowout counts as ~0.88 wins, while a
# 2-point squeaker counts as ~0.52 wins. The temperature parameter
# (20) controls how much margin matters vs. binary win/loss.

C = 2 * np.eye(n_teams)   # Colley matrix (diagonal = 2 + games played)
b = np.ones(n_teams)       # RHS vector (starts at 1 = prior of 0.5 per team)

for _, row in train.iterrows():
    hi = team_to_idx[row['HomeTeam']]
    ai = team_to_idx[row['AwayTeam']]
    margin = row['HomeWinMargin'] - HFA

    # Update Colley matrix (standard part)
    C[hi, hi] += 1; C[ai, ai] += 1
    C[hi, ai] -= 1; C[ai, hi] -= 1

    # Fractional win via sigmoid
    frac = 1 / (1 + np.exp(-margin / 20))
    b[hi] += (frac - 0.5)
    b[ai] += (0.5 - frac)

colley_ratings = np.linalg.solve(C, b)
colley_ratings -= colley_ratings.mean()
colley_dict = {all_teams[i]: colley_ratings[i] for i in range(n_teams)}


# ─── System 5: Feature-Based Rating ───────────────────────────────
# A weighted combination of four team-level statistics:
#   - Average neutral-site margin (50% weight) — overall strength
#   - Offensive output above league average (20%) — scoring ability
#   - Defensive output below league average (20%) — ability to limit opponents
#   - Win percentage deviation from .500 (10%) — clutch factor
# This captures dimensions that pure margin-based methods might miss.

team_stats = {}
for team in all_teams:
    gh = train[train['HomeTeam'] == team]
    ga = train[train['AwayTeam'] == team]

    pts_for     = list(gh['HomePts']) + list(ga['AwayPts'])
    pts_against = list(gh['AwayPts']) + list(ga['HomePts'])
    margins     = ([m - HFA for m in gh['HomeWinMargin']] +
                   [-(m - HFA) for m in ga['HomeWinMargin']])

    wins  = sum(1 for m in margins if m > 0)
    total = len(margins)

    team_stats[team] = {
        'avg_pf': np.mean(pts_for),
        'avg_pa': np.mean(pts_against),
        'avg_margin': np.mean(margins),
        'win_pct': wins / total,
    }

feat_rating = {}
for team in all_teams:
    s = team_stats[team]
    feat_rating[team] = (
        0.50 * s['avg_margin'] +                   # margin dominance
        0.20 * (s['avg_pf'] - 70) +                # offense above mean
        0.20 * (70 - s['avg_pa']) +                 # defense above mean
        0.10 * (s['win_pct'] - 0.5) * 100           # win% deviation
    )


# ════════════════════════════════════════════════════════════════════
# STEP 3: NORMALIZE ALL RATINGS TO A COMMON SCALE (Z-SCORES)
# ════════════════════════════════════════════════════════════════════
# Each system outputs ratings in different units (Elo points, raw
# margins, Colley values, etc). Z-score normalization puts them all
# on the same scale so weights are meaningful.

def normalize(ratings_dict):
    vals = np.array(list(ratings_dict.values()))
    mu, sigma = vals.mean(), vals.std()
    return {k: (v - mu) / sigma for k, v in ratings_dict.items()}

systems = {
    'massey': normalize(massey_dict),
    'elo':    normalize(elo_ratings),
    'adj_pd': normalize(adj_rating),
    'colley': normalize(colley_dict),
    'feat':   normalize(feat_rating),
}


# ════════════════════════════════════════════════════════════════════
# STEP 4: WEIGHTED ENSEMBLE
# ════════════════════════════════════════════════════════════════════
# Weights chosen to emphasize margin-based systems (Massey, Adj PD)
# which directly model the target variable, while giving meaningful
# weight to diverse systems (Elo, Colley, Feature-based) for
# robustness on unseen matchups.

weights = {
    'massey': 0.30,   # strongest individual predictor
    'adj_pd': 0.25,   # captures schedule strength
    'elo':    0.15,   # sequential, captures momentum/form
    'colley': 0.15,   # Bayesian, handles small samples well
    'feat':   0.15,   # captures offense/defense dimensions
}

ensemble = {}
for team in all_teams:
    ensemble[team] = sum(weights[s] * systems[s][team] for s in systems)


# ════════════════════════════════════════════════════════════════════
# STEP 5: CALIBRATE ENSEMBLE → POINT MARGIN
# ════════════════════════════════════════════════════════════════════
# The ensemble scores are in arbitrary z-score units. We fit a simple
# linear regression (OLS) to map score differences to actual margins:
#   predicted_margin = scale × (ensemble[team1] - ensemble[team2]) + intercept
# This learns the correct "exchange rate" between ensemble units and points.

train_diffs   = []
train_actuals = []

for _, row in train.iterrows():
    h, a = row['HomeTeam'], row['AwayTeam']
    train_diffs.append(ensemble[h] - ensemble[a])
    train_actuals.append(row['HomeWinMargin'] - HFA)  # neutral-site margin

train_diffs   = np.array(train_diffs)
train_actuals = np.array(train_actuals)

intercept, scale = polyfit(train_diffs, train_actuals, 1)
print(f"Calibration: margin = {scale:.2f} × ensemble_diff + {intercept:.2f}")


# ════════════════════════════════════════════════════════════════════
# STEP 6: APPLY SHRINKAGE (5%)
# ════════════════════════════════════════════════════════════════════
# With only ~11 games per team, ratings are noisy. Extreme predictions
# (large margins) are more likely to reflect noise than true team
# differences. A 5% shrinkage toward zero slightly dampens all
# predictions, which hurts training RMSE marginally but improves
# out-of-sample generalization.

SHRINKAGE = 0.95


# ════════════════════════════════════════════════════════════════════
# STEP 7: GENERATE DERBY PREDICTIONS
# ════════════════════════════════════════════════════════════════════
# For each derby match (neutral site):
#   predicted_margin = shrinkage × (scale × (rating[t1] - rating[t2]) + intercept)
# Positive = Team1 wins, Negative = Team2 wins.

pred_output = pred.copy()

for i, row in pred_output.iterrows():
    t1, t2 = row['Team1'], row['Team2']
    raw_diff = ensemble[t1] - ensemble[t2]
    margin = (scale * raw_diff + intercept) * SHRINKAGE
    pred_output.at[i, 'Team1_WinMargin'] = round(margin)

pred_output['Team1_WinMargin'] = pred_output['Team1_WinMargin'].astype(int)
pred_output.to_csv('Predictions_output.csv', index=False)

# Print diagnostics
margins = pred_output['Team1_WinMargin']
print(f"\nDerby prediction stats:")
print(f"  Mean margin: {margins.mean():.1f}")
print(f"  Std dev:     {margins.std():.1f}")
print(f"  Range:       [{margins.min()}, {margins.max()}]")
print(f"  Team1 wins:  {(margins > 0).sum()}")
print(f"  Team2 wins:  {(margins < 0).sum()}")
print(f"  Ties:        {(margins == 0).sum()}")


# ════════════════════════════════════════════════════════════════════
# STEP 8: GENERATE TEAM RANKINGS (1-165)
# ════════════════════════════════════════════════════════════════════
# Rank teams by their ensemble rating (higher = better).
# This uses the same ensemble that drives predictions, ensuring
# consistency between predicted margins and team rankings.

rank_order = sorted(ensemble.items(), key=lambda x: -x[1])
team_rank = {t: r for r, (t, _) in enumerate(rank_order, 1)}

wb = openpyxl.load_workbook('Rankings.xlsx')
ws = wb.active
for ri in range(2, ws.max_row + 1):
    tn = ws.cell(row=ri, column=2).value
    if tn in team_rank:
        ws.cell(row=ri, column=3).value = team_rank[tn]
wb.save('Rankings_output.xlsx')

print(f"\nTop 10 Rankings:")
for r, (t, s) in enumerate(rank_order[:10], 1):
    print(f"  {r:2d}. {t}")


# ════════════════════════════════════════════════════════════════════
# TRAINING DIAGNOSTICS
# ════════════════════════════════════════════════════════════════════

cal_preds = (scale * train_diffs + intercept) * SHRINKAGE
rmse = np.sqrt(np.mean((cal_preds - train_actuals) ** 2))
correct = sum(1 for p, a in zip(cal_preds, train_actuals) if (p > 0) == (a > 0))

print(f"\nTraining performance:")
print(f"  RMSE:           {rmse:.2f}")
print(f"  Correct winner: {correct}/{len(train_actuals)} ({correct/len(train_actuals):.1%})")
